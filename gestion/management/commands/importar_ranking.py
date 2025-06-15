from django.core.management.base import BaseCommand
from gestion.models import Atleta, Entrenador
import openpyxl
from django.utils.text import slugify
from decimal import Decimal, InvalidOperation
from datetime import date

class Command(BaseCommand):
    help = 'Importa atletas y puntos de ranking desde un archivo Excel con formato personalizado'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta del archivo Excel a importar')

    def handle(self, *args, **kwargs):
        excel_file = kwargs['excel_file']
        wb = openpyxl.load_workbook(excel_file, data_only=True)  # data_only=True = ignora fórmulas, usa valores

        try:
            entrenador_defecto = Entrenador.objects.get(nombre_completo="IMPORTADO")
        except Entrenador.DoesNotExist:
            self.stdout.write(self.style.ERROR("❌ No se encontró el entrenador 'IMPORTADO'"))
            return

        total_creados = 0
        total_actualizados = 0

        for hoja in wb.sheetnames:
            ws = wb[hoja]
            self.stdout.write(f"\n📄 Procesando hoja: {hoja}")

            for row in ws.iter_rows(min_row=6, values_only=True):  # Desde fila 6
                if not row or len(row) < 6:
                    continue

                nombre = row[2] or ""
                gimnasio = row[3] or "SIN GIMNASIO"
                puntos_2025 = row[5]

                if not nombre.strip() or not puntos_2025:
                    continue

                try:
                    puntos = Decimal(puntos_2025)
                except (InvalidOperation, TypeError):
                    continue

                cui_temp = f"TEMP_{slugify(nombre)}_{slugify(gimnasio)}"[:20]

                atleta, creado = Atleta.objects.update_or_create(
                    cui=cui_temp,
                    defaults={
                        "nombre_completo": nombre.strip(),
                        "genero": "femenino" if "FEMENINO" in hoja.upper() else "masculino",
                        "categoria": "cadete" if "CADETE" in hoja.upper()
                                     else "juvenil" if "JUVENIL" in hoja.upper()
                                     else "adulto" if "ADULTO" in hoja.upper()
                                     else None,
                        "disciplina": "poomsae" if "POOMSAE" in hoja.upper() or "FREESTYLE" in hoja.upper() else "combate",
                        "peso": "",  # aún sin definir
                        "puntos_ranking": puntos,
                        "entrenador": entrenador_defecto,
                        "cinta": "roja",
                        "fecha_nacimiento": date(2000, 1, 1),
                    }
                )

                if creado:
                    total_creados += 1
                else:
                    total_actualizados += 1

        self.stdout.write(self.style.SUCCESS("\n✅ Importación completada"))
        self.stdout.write(f"Atletas creados: {total_creados}")
        self.stdout.write(f"Atletas actualizados: {total_actualizados}")

